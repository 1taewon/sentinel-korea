"""process_kdca_excel.py - KDCA XLSX/CSV/PDF ingestion for Sentinel Korea.

Supported user upload bundle:
- 인플루엔자 선택됨 - 17wk.xlsx
- 급성호흡기감염증 선택됨- 전체[전체] - 17wk.xlsx
- 중증급성호흡기감염증 선택됨 환자 중 폐렴 환자 현황 - 17wk.xlsx
- 중증급성호흡기감염증 선택됨 환자 중 인플루엔자 환자 현황 - 17wk.xlsx
- 하수기반감염병감시주간분석보고 PDF
- 감염병 표본감시 주간소식지 PDF
- 전 세계 감염병 발생 동향 PDF
"""
from __future__ import annotations

import csv
import json
import os
import re
import sys
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

PROCESSED_DIR = Path(__file__).resolve().parent.parent / "data" / "processed"
SNAPSHOT_DIR = PROCESSED_DIR / "snapshots"
UPLOAD_HISTORY_FILE = PROCESSED_DIR / "upload_history.json"

SENTINEL_DATA_DIR = Path(
    os.getenv("SENTINEL_DATA_DIR", r"C:\Users\han75\OneDrive\Desktop\Sentinel_data")
)

STRUCTURED_OUTPUT_FILES = {
    "ari": "kdca_ari_weekly.json",
    "influenza": "kdca_influenza_ili_weekly.json",
    "sari_pneumonia": "kdca_sari_pneumonia_weekly.json",
    "sari_influenza": "kdca_sari_influenza_weekly.json",
}

FILE_TYPE_LABELS = {
    "ari": "급성호흡기감염증 전체",
    "influenza": "인플루엔자 의사환자분율",
    "sari_pneumonia": "중증급성호흡기감염증 중 폐렴 환자",
    "sari_influenza": "중증급성호흡기감염증 중 인플루엔자 환자",
    "wastewater_pdf": "하수기반감염병감시 주간분석보고",
    "surveillance_bulletin_pdf": "감염병 표본감시 주간소식지",
    "global_outbreak_pdf": "전 세계 감염병 발생 동향",
}

# Order matters: SARI-specific filenames also contain "인플루엔자".
PATTERNS: list[tuple[str, re.Pattern[str]]] = [
    ("sari_influenza", re.compile(r"중증급성.*인플루엔자|SARI.*influenza", re.IGNORECASE)),
    ("sari_pneumonia", re.compile(r"중증급성.*폐렴|SARI.*pneumonia", re.IGNORECASE)),
    ("ari", re.compile(r"급성호흡기감염증|acute.*respiratory|ARI", re.IGNORECASE)),
    ("influenza", re.compile(r"^.*인플루엔자\s*선택됨|influenza|ILI", re.IGNORECASE)),
    ("wastewater_pdf", re.compile(r"하수기반|wastewater", re.IGNORECASE)),
    ("surveillance_bulletin_pdf", re.compile(r"표본감시\s*주간소식지|weekly.*surveillance", re.IGNORECASE)),
    ("global_outbreak_pdf", re.compile(r"전\s*세계\s*감염병\s*발생\s*동향|global.*outbreak", re.IGNORECASE)),
]

COUNTRY_GEO = {
    "홍콩": {"country": "Hong Kong", "lat": 22.3193, "lng": 114.1694},
    "미국": {"country": "United States", "lat": 39.8283, "lng": -98.5795},
    "방글라데시": {"country": "Bangladesh", "lat": 23.685, "lng": 90.3563},
    "중국": {"country": "China", "lat": 35.8617, "lng": 104.1954},
    "일본": {"country": "Japan", "lat": 36.2048, "lng": 138.2529},
    "베트남": {"country": "Vietnam", "lat": 14.0583, "lng": 108.2772},
    "태국": {"country": "Thailand", "lat": 15.87, "lng": 100.9925},
    "필리핀": {"country": "Philippines", "lat": 12.8797, "lng": 121.774},
    "인도네시아": {"country": "Indonesia", "lat": -0.7893, "lng": 113.9213},
    "말레이시아": {"country": "Malaysia", "lat": 4.2105, "lng": 101.9758},
    "인도": {"country": "India", "lat": 20.5937, "lng": 78.9629},
}


def epiweek_to_date(year: int, week: int) -> str:
    """ISO epiweek -> Sunday date used by the current snapshot calendar."""
    jan4 = date(year, 1, 4)
    week1_monday = jan4 - timedelta(days=jan4.weekday())
    target = week1_monday + timedelta(weeks=week - 1, days=6)
    return target.strftime("%Y-%m-%d")


def detect_file_type(filename: str) -> str | None:
    for ftype, pattern in PATTERNS:
        if pattern.search(filename):
            return ftype
    return None


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "None", "집계 중", "집계중"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_int(value: Any) -> int | None:
    numeric = parse_number(value)
    if numeric is None:
        return None
    return int(numeric)


def epiweek(year: int, week: int) -> str:
    return f"{year}-W{week:02d}"


def normalize_signal(value: float, min_val: float, max_val: float) -> float:
    if max_val == min_val:
        return 0.5
    return max(0.0, min(1.0, (value - min_val) / (max_val - min_val)))


def value_range(records: list[dict[str, Any]], key: str) -> tuple[float, float]:
    values = [float(r[key]) for r in records if r.get(key) is not None]
    if not values:
        return 0.0, 1.0
    return min(values), max(values)


def latest_completed(records: list[dict[str, Any]], value_key: str) -> dict[str, Any] | None:
    completed = [r for r in records if r.get(value_key) is not None]
    if not completed:
        return None
    return sorted(completed, key=lambda r: (int(r["year"]), int(r["week"])))[-1]


def save_json(filename: str, payload: Any) -> str:
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    path = PROCESSED_DIR / filename
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return filename


def update_surveillance_catalog(file_type: str, filename: str, records: list[dict[str, Any]], output_file: str) -> None:
    catalog_path = PROCESSED_DIR / "kdca_uploaded_surveillance_catalog.json"
    if catalog_path.exists():
        try:
            catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
        except Exception:
            catalog = {}
    else:
        catalog = {}

    catalog.setdefault("status", "ok")
    catalog.setdefault("sources", {})
    completed = [r for r in records if not r.get("pending")]
    catalog["updated_at"] = datetime.utcnow().isoformat()
    catalog["sources"][file_type] = {
        "label": FILE_TYPE_LABELS.get(file_type, file_type),
        "filename": filename,
        "output_file": output_file,
        "records": len(records),
        "completed_records": len(completed),
        "latest_completed_epiweek": latest_completed(records, "value") and latest_completed(records, "value")["epiweek"],
    }
    catalog_path.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")


def save_structured_source(
    file_type: str,
    filename: str,
    records: list[dict[str, Any]],
    value_key: str,
    extra: dict[str, Any] | None = None,
) -> str:
    completed = [r for r in records if r.get(value_key) is not None]
    pending = [r for r in records if r.get(value_key) is None]
    latest = latest_completed(records, value_key)
    payload = {
        "status": "ok",
        "source_type": file_type,
        "label": FILE_TYPE_LABELS.get(file_type, file_type),
        "filename": filename,
        "scope": "national_weekly",
        "completed_records": len(completed),
        "pending_records": len(pending),
        "latest_completed_epiweek": latest["epiweek"] if latest else None,
        "latest_completed_period": f"{latest['year']}년 {latest['week']:02d}주" if latest else None,
        "records": records,
    }
    if extra:
        payload.update(extra)
    output = STRUCTURED_OUTPUT_FILES[file_type]
    save_json(output, payload)
    update_surveillance_catalog(file_type, filename, records, output)
    return output


def parse_ari_csv(content: str) -> list[dict[str, Any]]:
    records = []
    reader = csv.reader(StringIO(content))
    for row in reader:
        if not row or not row[0].strip().isdigit():
            continue
        year = parse_int(row[0])
        week = parse_int(row[1] if len(row) > 1 else None)
        total = parse_number(row[2] if len(row) > 2 else None)
        if not year or not week:
            continue
        records.append({
            "year": year,
            "week": week,
            "epiweek": epiweek(year, week),
            "date": epiweek_to_date(year, week),
            "value": total,
            "total": total,
            "pending": total is None,
        })
    return records


def parse_influenza_csv(content: str) -> list[dict[str, Any]]:
    records = []
    reader = csv.reader(StringIO(content))
    for row in reader:
        if not row:
            continue
        season_match = re.search(r"(\d{4})-(\d{4})", row[0])
        if not season_match:
            continue
        start_year = int(season_match.group(1))
        end_year = int(season_match.group(2))
        for idx, cell in enumerate(row[1:], start=36):
            week = idx
            actual_year = start_year if week >= 36 else end_year
            if week > 52:
                week -= 52
                actual_year = end_year
            value = parse_number(cell)
            records.append({
                "year": actual_year,
                "week": week,
                "epiweek": epiweek(actual_year, week),
                "date": epiweek_to_date(actual_year, week),
                "value": value,
                "ili_index": value,
                "pending": value is None,
            })
    return records


def parse_sari_csv(content: str, file_type: str = "sari_pneumonia") -> list[dict[str, Any]]:
    records = []
    reader = csv.reader(StringIO(content))
    for row in reader:
        values = [parse_number(cell) for cell in row]
        numeric = [v for v in values if v is not None]
        if not numeric:
            continue
        for idx, val in enumerate(numeric):
            week = 36 + idx
            year = date.today().year - 1 if week >= 36 else date.today().year
            if week > 52:
                week -= 52
                year = date.today().year
            metric = "sari_influenza_cases" if file_type == "sari_influenza" else "sari_pneumonia_cases"
            records.append({
                "year": year,
                "week": week,
                "epiweek": epiweek(year, week),
                "date": epiweek_to_date(year, week),
                "value": val,
                metric: val,
                "pending": False,
            })
        break
    return records


def parse_ari_xlsx(ws: Any) -> list[dict[str, Any]]:
    pathogen_columns = {
        5: "마이코플라즈마균",
        6: "클라미디아균",
        7: "아데노바이러스",
        8: "사람 보카바이러스",
        9: "파라인플루엔자바이러스",
        10: "호흡기세포융합바이러스",
        11: "리노바이러스",
        12: "사람 메타뉴모바이러스",
        13: "사람 코로나바이러스",
        14: "인플루엔자 바이러스",
        15: "코로나19 바이러스",
    }
    records: list[dict[str, Any]] = []
    current_year: int | None = None
    for row_idx in range(9, ws.max_row + 1):
        year_val = parse_int(ws.cell(row_idx, 2).value)
        if year_val:
            current_year = year_val
        week = parse_int(ws.cell(row_idx, 3).value)
        if not current_year or not week:
            continue
        total = parse_number(ws.cell(row_idx, 4).value)
        pathogens = {
            pathogen: parse_number(ws.cell(row_idx, col).value)
            for col, pathogen in pathogen_columns.items()
        }
        records.append({
            "year": current_year,
            "week": week,
            "epiweek": epiweek(current_year, week),
            "date": epiweek_to_date(current_year, week),
            "value": total,
            "total": total,
            "pathogens": pathogens,
            "pending": total is None,
        })
    return records


def parse_influenza_xlsx(ws: Any) -> list[dict[str, Any]]:
    season_text = str(ws.cell(8, 2).value or "")
    season = re.search(r"(\d{4})-(\d{4})", season_text)
    start_year = int(season.group(1)) if season else date.today().year - 1
    end_year = int(season.group(2)) if season else date.today().year
    records: list[dict[str, Any]] = []
    for col in range(3, ws.max_column + 1):
        week = parse_int(str(ws.cell(7, col).value or "").replace("주", ""))
        if not week:
            continue
        actual_year = start_year if week >= 36 else end_year
        value = parse_number(ws.cell(8, col).value)
        records.append({
            "year": actual_year,
            "week": week,
            "epiweek": epiweek(actual_year, week),
            "date": epiweek_to_date(actual_year, week),
            "value": value,
            "ili_index": value,
            "pending": value is None,
        })
    return records


def parse_sari_xlsx(ws: Any, file_type: str) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    current_year: int | None = None
    metric_key = "sari_influenza_cases" if file_type == "sari_influenza" else "sari_pneumonia_cases"
    for col in range(3, ws.max_column + 1):
        year_val = parse_int(ws.cell(7, col).value)
        if year_val:
            current_year = year_val
        week = parse_int(ws.cell(8, col).value)
        if not current_year or not week:
            continue
        total = parse_number(ws.cell(9, col).value)
        selected = parse_number(ws.cell(10, col).value)
        ratio = None if total in (None, 0) or selected is None else round(selected / total, 4)
        records.append({
            "year": current_year,
            "week": week,
            "epiweek": epiweek(current_year, week),
            "date": epiweek_to_date(current_year, week),
            "value": selected,
            "sari_total": total,
            metric_key: selected,
            "case_ratio": ratio,
            "pending": selected is None,
        })
    return records


def build_snapshot_from_records(
    ari_records: list[dict[str, Any]],
    ili_records: list[dict[str, Any]],
    sari_pneumonia_records: list[dict[str, Any]],
    sari_influenza_records: list[dict[str, Any]],
) -> dict[str, dict[str, float]]:
    date_signals: dict[str, dict[str, float]] = {}

    ari_min, ari_max = value_range(ari_records, "total")
    for record in ari_records:
        if record.get("total") is None:
            continue
        date_signals.setdefault(record["date"], {})["notifiable_disease"] = round(
            normalize_signal(float(record["total"]), ari_min, ari_max), 4
        )

    ili_min, ili_max = value_range(ili_records, "ili_index")
    for record in ili_records:
        if record.get("ili_index") is None:
            continue
        date_signals.setdefault(record["date"], {})["influenza_like"] = round(
            normalize_signal(float(record["ili_index"]), ili_min, ili_max), 4
        )

    pneu_min, pneu_max = value_range(sari_pneumonia_records, "value")
    for record in sari_pneumonia_records:
        if record.get("value") is None:
            continue
        date_signals.setdefault(record["date"], {})["sari_pneumonia"] = round(
            normalize_signal(float(record["value"]), pneu_min, pneu_max), 4
        )

    flu_min, flu_max = value_range(sari_influenza_records, "value")
    for record in sari_influenza_records:
        if record.get("value") is None:
            continue
        date_signals.setdefault(record["date"], {})["sari_influenza"] = round(
            normalize_signal(float(record["value"]), flu_min, flu_max), 4
        )

    return date_signals


def update_snapshots(date_signals: dict[str, dict[str, float]]) -> list[str]:
    """Apply national weekly signals to snapshots without inventing regional variation."""
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    updated_dates: list[str] = []

    for snapshot_date, signals in sorted(date_signals.items()):
        snapshot_path = SNAPSHOT_DIR / f"{snapshot_date}.json"

        if snapshot_path.exists():
            snapshot_data = json.loads(snapshot_path.read_text(encoding="utf-8"))
        else:
            existing = sorted(SNAPSHOT_DIR.glob("*.json"))
            if not existing:
                continue
            import copy
            snapshot_data = copy.deepcopy(json.loads(existing[-1].read_text(encoding="utf-8")))
            for region in snapshot_data:
                region["date"] = snapshot_date

        for region in snapshot_data:
            region.setdefault("signals", {})
            region.setdefault("signal_scope", {})
            for sig_key, sig_val in signals.items():
                region["signals"][sig_key] = sig_val
                region["signal_scope"][sig_key] = "national_uploaded_kdca"

        snapshot_path.write_text(
            json.dumps(snapshot_data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        updated_dates.append(snapshot_date)

    return updated_dates


def record_upload_history(
    filename: str,
    file_type: str,
    updated_dates: list[str],
    outputs: list[str] | None = None,
    records_parsed: int = 0,
) -> None:
    history = []
    if UPLOAD_HISTORY_FILE.exists():
        try:
            history = json.loads(UPLOAD_HISTORY_FILE.read_text(encoding="utf-8"))
        except Exception:
            history = []

    history.append({
        "filename": filename,
        "file_type": file_type,
        "label": FILE_TYPE_LABELS.get(file_type, file_type),
        "uploaded_at": datetime.utcnow().isoformat(),
        "updated_dates": updated_dates,
        "snapshot_count": len(updated_dates),
        "records_parsed": records_parsed,
        "outputs": outputs or [],
    })
    history = history[-50:]
    UPLOAD_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")


def _process_xlsx(filename: str, file_type: str, raw_bytes: bytes) -> dict[str, Any]:
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(raw_bytes), data_only=True)
    ws = wb.active
    if file_type == "ari":
        records = parse_ari_xlsx(ws)
        output = save_structured_source(file_type, filename, records, "total")
        date_signals = build_snapshot_from_records(records, [], [], [])
    elif file_type == "influenza":
        records = parse_influenza_xlsx(ws)
        output = save_structured_source(file_type, filename, records, "ili_index")
        date_signals = build_snapshot_from_records([], records, [], [])
    elif file_type == "sari_pneumonia":
        records = parse_sari_xlsx(ws, file_type)
        output = save_structured_source(file_type, filename, records, "value")
        date_signals = build_snapshot_from_records([], [], records, [])
    elif file_type == "sari_influenza":
        records = parse_sari_xlsx(ws, file_type)
        output = save_structured_source(file_type, filename, records, "value")
        date_signals = build_snapshot_from_records([], [], [], records)
    else:
        raise ValueError(f"Unsupported XLSX file type: {file_type}")

    updated_dates = update_snapshots(date_signals)
    record_upload_history(filename, file_type, updated_dates, [output], len(records))
    return {
        "success": True,
        "filename": filename,
        "file_type": file_type,
        "label": FILE_TYPE_LABELS[file_type],
        "records_parsed": len(records),
        "snapshots_updated": len(updated_dates),
        "updated_dates": updated_dates,
        "outputs": [output],
    }


def extract_pdf_pages(raw_bytes: bytes) -> list[dict[str, Any]]:
    import pdfplumber

    pages = []
    with pdfplumber.open(BytesIO(raw_bytes)) as pdf:
        for idx, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            try:
                tables = page.extract_tables()
            except Exception:
                tables = []
            pages.append({
                "page": idx,
                "text": text,
                "tables": tables,
                "table_count": len(tables),
            })
    return pages


def parse_wastewater_pdf(filename: str, pages: list[dict[str, Any]]) -> tuple[dict[str, Any], str]:
    all_text = "\n".join(page["text"] for page in pages)
    week_match = re.search(r"(\d{4})년도\s*(\d+)주차\(([^)]+)\)", all_text)
    facility_match = re.search(r"하수처리장\((\d+)개소\)", all_text)
    pathogens = []
    for pathogen in ["SARS-CoV-2", "Influenza virus", "Norovirus"]:
        if pathogen in all_text:
            pathogens.append(pathogen)

    region_sections = []
    for page in pages:
        title_match = re.search(r"q\s*(.+?지역별 하수 기반 감염병 감시 현황)", page["text"])
        if title_match:
            region_sections.append({
                "page": page["page"],
                "section": title_match.group(1),
                "review_required": True,
            })

    payload = {
        "status": "ok",
        "source_type": "wastewater_pdf",
        "filename": filename,
        "year": int(week_match.group(1)) if week_match else None,
        "week": int(week_match.group(2)) if week_match else None,
        "period": week_match.group(3) if week_match else None,
        "facilities": int(facility_match.group(1)) if facility_match else None,
        "pathogens": pathogens,
        "extraction_mode": "text_metadata_only",
        "data_quality": "needs_review",
        "note": "지역별 농도/범주 차트는 PDF 이미지로 제공되어 자동 수치화는 검수 모드에서 처리해야 합니다.",
        "region_sections": region_sections,
        "page_count": len(pages),
        "text_excerpt": all_text[:2200],
    }
    output = save_json("kdca_wastewater_pdf_evidence.json", payload)
    return payload, output


def infer_severity(disease: str, text: str) -> str:
    combined = f"{disease} {text}".lower()
    if any(token in combined for token in ["홍역", "measles", "mers", "sars", "avian", "h5n1", "사망"]):
        return "high"
    if any(token in combined for token in ["뎅기", "dengue", "salmonella", "outbreak", "발생"]):
        return "medium"
    return "low"


def parse_global_outbreak_pdf(filename: str, pages: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], str]:
    all_text = "\n".join(page["text"] for page in pages)
    date_match = re.search(r"(\d{4})년\s*제\d+호\((\d{4})\.(\d{1,2})\.(\d{1,2})\.\)", all_text)
    report_date = None
    if date_match:
        report_date = f"{date_match.group(2)}-{int(date_match.group(3)):02d}-{int(date_match.group(4)):02d}"

    heading_candidates: dict[tuple[str, str], str] = {}
    for page in pages:
        for line in page["text"].splitlines():
            line = line.strip()
            match = re.match(r"^(\d+)\.\s*([^,]+),\s*([^\s]+)\s+(.+)$", line)
            if not match:
                continue
            disease_kr = match.group(2).strip()
            country_kr = match.group(3).strip()
            key = (disease_kr, country_kr)
            if key not in heading_candidates or len(line) > len(heading_candidates[key]):
                heading_candidates[key] = line

    signals: list[dict[str, Any]] = []
    for idx, ((disease_kr, country_kr), heading) in enumerate(heading_candidates.items(), start=1):
        geo = COUNTRY_GEO.get(country_kr, {"country": country_kr, "lat": 0.0, "lng": 0.0})
        snippet_start = all_text.find(heading)
        snippet = all_text[snippet_start:snippet_start + 900] if snippet_start >= 0 else heading
        severity = infer_severity(disease_kr, snippet)
        signals.append({
            "id": f"kdca-global-{idx:02d}",
            "source": "kdca_global_report",
            "title": heading,
            "lat": geo["lat"],
            "lng": geo["lng"],
            "date": report_date or datetime.utcnow().date().isoformat(),
            "disease": disease_kr,
            "country": geo["country"],
            "severity": severity,
            "url": "",
            "raw_excerpt": snippet,
        })

    output = save_json("global_kdca_outbreaks.json", signals)
    evidence = {
        "status": "ok",
        "source_type": "global_outbreak_pdf",
        "filename": filename,
        "report_date": report_date,
        "signals_file": output,
        "signals": signals,
        "page_count": len(pages),
        "text_excerpt": all_text[:2500],
    }
    save_json("kdca_global_outbreak_pdf_evidence.json", evidence)
    return signals, output


def parse_surveillance_bulletin_pdf(filename: str, pages: list[dict[str, Any]]) -> tuple[dict[str, Any], str]:
    all_text = "\n".join(page["text"] for page in pages)
    week_match = re.search(r"(\d{4})년도.*?(\d{1,2})주차", filename + "\n" + all_text)
    payload = {
        "status": "ok",
        "source_type": "surveillance_bulletin_pdf",
        "filename": filename,
        "year": int(week_match.group(1)) if week_match else None,
        "week": int(week_match.group(2)) if week_match else None,
        "extraction_mode": "text_evidence",
        "note": "이 PDF는 구조화 xlsx 3종을 질병관리청이 보고서로 편집한 문서 evidence로 저장합니다.",
        "page_count": len(pages),
        "text_excerpt": all_text[:4000],
    }
    output = save_json("kdca_weekly_bulletin_pdf_evidence.json", payload)
    return payload, output


def _process_pdf(filename: str, file_type: str, raw_bytes: bytes) -> dict[str, Any]:
    pages = extract_pdf_pages(raw_bytes)
    if file_type == "wastewater_pdf":
        payload, output = parse_wastewater_pdf(filename, pages)
        records_parsed = len(payload.get("region_sections", []))
    elif file_type == "global_outbreak_pdf":
        signals, output = parse_global_outbreak_pdf(filename, pages)
        records_parsed = len(signals)
    elif file_type == "surveillance_bulletin_pdf":
        payload, output = parse_surveillance_bulletin_pdf(filename, pages)
        records_parsed = 1 if payload else 0
    else:
        raise ValueError(f"Unsupported PDF file type: {file_type}")

    record_upload_history(filename, file_type, [], [output], records_parsed)
    return {
        "success": True,
        "filename": filename,
        "file_type": file_type,
        "label": FILE_TYPE_LABELS[file_type],
        "records_parsed": records_parsed,
        "snapshots_updated": 0,
        "updated_dates": [],
        "outputs": [output],
    }


def process_file(file_path: Path, content_bytes: bytes | None = None) -> dict[str, Any]:
    filename = file_path.name
    file_type = detect_file_type(filename)
    if not file_type:
        return {"success": False, "filename": filename, "error": f"알 수 없는 파일 형식: {filename}"}

    try:
        raw_bytes = content_bytes if content_bytes is not None else file_path.read_bytes()
        suffix = file_path.suffix.lower()
        if suffix == ".xlsx":
            return _process_xlsx(filename, file_type, raw_bytes)
        if suffix == ".pdf":
            return _process_pdf(filename, file_type, raw_bytes)

        for encoding in ["utf-8-sig", "utf-8", "cp949", "euc-kr"]:
            try:
                content = raw_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            return {"success": False, "filename": filename, "error": "인코딩 변환 실패"}

        if file_type == "ari":
            ari_records = parse_ari_csv(content)
            output = save_structured_source(file_type, filename, ari_records, "total")
            date_signals = build_snapshot_from_records(ari_records, [], [], [])
            records = ari_records
        elif file_type == "influenza":
            ili_records = parse_influenza_csv(content)
            output = save_structured_source(file_type, filename, ili_records, "ili_index")
            date_signals = build_snapshot_from_records([], ili_records, [], [])
            records = ili_records
        elif file_type in {"sari_pneumonia", "sari_influenza"}:
            sari_records = parse_sari_csv(content, file_type)
            output = save_structured_source(file_type, filename, sari_records, "value")
            date_signals = build_snapshot_from_records(
                [], [], sari_records if file_type == "sari_pneumonia" else [], sari_records if file_type == "sari_influenza" else []
            )
            records = sari_records
        else:
            return {"success": False, "filename": filename, "error": f"CSV로 처리할 수 없는 파일 형식: {file_type}"}

        updated_dates = update_snapshots(date_signals)
        record_upload_history(filename, file_type, updated_dates, [output], len(records))
        return {
            "success": True,
            "filename": filename,
            "file_type": file_type,
            "label": FILE_TYPE_LABELS[file_type],
            "records_parsed": len(records),
            "snapshots_updated": len(updated_dates),
            "updated_dates": updated_dates,
            "outputs": [output],
        }
    except Exception as e:
        return {"success": False, "filename": filename, "file_type": file_type, "error": str(e)}


def scan_sentinel_data_dir() -> list[dict[str, Any]]:
    """Recursively scan SENTINEL_DATA_DIR for KDCA files.

    Walks every subdirectory (e.g. 17wk/, 18wk/) so the user can keep
    weekly downloads grouped by epi-week without flattening.
    """
    if not SENTINEL_DATA_DIR.exists():
        print(f"[KDCA] 데이터 폴더 없음: {SENTINEL_DATA_DIR}", file=sys.stderr)
        return []

    processed_files = set()
    if UPLOAD_HISTORY_FILE.exists():
        try:
            history = json.loads(UPLOAD_HISTORY_FILE.read_text(encoding="utf-8"))
            processed_files = {h["filename"] for h in history}
        except Exception:
            pass

    results: list[dict[str, Any]] = []
    for ext in ["*.csv", "*.xlsx", "*.pdf"]:
        # rglob = recursive — picks up files under 17wk/, 18wk/, ...
        for file_path in SENTINEL_DATA_DIR.rglob(ext):
            if file_path.name in processed_files:
                continue
            if detect_file_type(file_path.name) is None:
                continue
            print(f"[KDCA] 처리 중: {file_path.relative_to(SENTINEL_DATA_DIR)}")
            result = process_file(file_path)
            results.append(result)
            if result.get("success"):
                print(f"[KDCA] 완료: {result.get('records_parsed', 0)} records, {result.get('snapshots_updated', 0)} snapshots")
            else:
                print(f"[KDCA] 실패: {result.get('error')}", file=sys.stderr)

    return results


if __name__ == "__main__":
    import dotenv

    dotenv.load_dotenv(Path(__file__).parent.parent / ".env")
    results = scan_sentinel_data_dir()
    print(f"\n처리 완료: {len(results)}개 파일")
    for result in results:
        print(f"  - {result.get('filename')}: {'성공' if result.get('success') else '실패'}")
